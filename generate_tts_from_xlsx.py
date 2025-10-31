#!/usr/bin/env python3
import argparse
import asyncio
import os
import re
import shlex
import subprocess
import sys
import tempfile
from pathlib import Path
import math
import time
import json
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional

import openpyxl  # type: ignore


def run(cmd: str) -> None:
    proc = subprocess.run(cmd, shell=True)
    if proc.returncode != 0:
        raise RuntimeError(f"Command failed: {cmd}")


def sanitize_filename(name: str) -> str:
    name = name.strip()
    # Replace invalid filename characters
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    # Collapse spaces/underscores
    name = re.sub(r"[\s_]+", "_", name)
    return name[:180] or "untitled"


def extract_product_and_date(filename: str) -> tuple[str, str]:
    """从文件名提取产品名称和日期"""
    name_without_ext = Path(filename).stem
    
    # 提取日期（YYYY-MM-DD 或 YYYYMMDD）
    date_match = re.search(r'(\d{4})[-_]?(\d{2})[-_]?(\d{2})', name_without_ext)
    date_str = ""
    if date_match:
        date_str = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
    
    # 提取产品名称（移除日期、数字后缀、合并/模板等）
    patterns = [
        (r'\d{4}[-_]?\d{2}[-_]?\d{2}[-_](.*?)_\d+', 1),  # 日期_产品名_数字
        (r'\d{4}[-_]?\d{2}[-_]?\d{2}[-_](.*?)_合并', 1),  # 日期_产品名_合并
        (r'\d{4}[-_]?\d{2}[-_]?\d{2}[-_](.*?)_模板', 1),  # 日期_产品名_模板
        (r'(.*?)_\d{4}[-_]?\d{2}[-_]?\d{2}', 1),      # 产品名_日期
        (r'(.*?)_\d+$', 1),                           # 产品名_数字
        (r'(.*?)_合并$', 1),                          # 产品名_合并
        (r'(.*?)_模板$', 1),                          # 产品名_模板
        (r'(.*?)_GPT$', 1),                           # 产品名_GPT
        (r'(.*?)_AI$', 1),                            # 产品名_AI
        (r'(.*?)_生成$', 1),                          # 产品名_生成
        (r'(.*?)_全产品.*?版', 1),                    # 全产品_合并版_3200_v1 类
    ]
    
    product_name = name_without_ext  # 默认使用整个文件名
    for pattern, group_idx in patterns:
        match = re.search(pattern, name_without_ext)
        if match:
            product_name = match.group(group_idx).strip()
            break
    
    # 清理产品名中的常见后缀
    product_name = re.sub(r'[_全产品合并版\d+_v\d+]+$', '', product_name).strip()
    
    # 如果没有日期，使用当前日期
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")
    
    return sanitize_filename(product_name), date_str


def split_into_sentences(text: str) -> list[str]:
    # Split by common sentence-ending punctuation; keep delimiters
    parts = re.split(r"([.!?。！？]|؟)", text)
    sentences: list[str] = []
    buf = ""
    for i in range(0, len(parts), 2):
        chunk = parts[i].strip()
        delim = parts[i + 1] if i + 1 < len(parts) else ""
        if not chunk and not delim:
            continue
        candidate = (chunk + (delim or "")).strip()
        if candidate:
            sentences.append(candidate)
    if not sentences and text.strip():
        sentences = [text.strip()]
    return sentences


async def tts_segment_async(segment_text: str, voice: str, rate: str, out_m4a: Path, max_retries: int = 5) -> bool:
    """异步 TTS 生成（edge-tts 默认格式，自动转换为 M4A），自动重试"""
    import edge_tts
    temp_file = out_m4a.with_suffix('.tmp')
    for attempt in range(max_retries):
        try:
            communicate = edge_tts.Communicate(segment_text, voice=voice, rate=rate)
            await communicate.save(str(temp_file))
            # edge-tts 默认可能输出 mp3/webm，统一转换为 M4A
            cmd = f"ffmpeg -threads 0 -hide_banner -loglevel error -i {shlex.quote(str(temp_file))} -c:a aac -b:a 192k {shlex.quote(str(out_m4a))} -y"
            run(cmd)
            if temp_file.exists():
                temp_file.unlink()
            return True
        except Exception as e:
            if attempt < max_retries - 1:
                await asyncio.sleep(min(2 ** attempt, 10))
                continue
            if temp_file.exists():
                temp_file.unlink()
            raise
    return False


def tts_segment(segment_text: str, voice: str, rate: str, volume: str, pitch: str, out_m4a: Path) -> None:
    """同步包装，内部调用异步版本"""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(tts_segment_async(segment_text, voice, rate, out_m4a))
    finally:
        loop.close()


def make_silence_m4a(duration_ms: int, out_m4a: Path, sample_rate: int = 24000) -> None:
    """生成 M4A 格式静音片段（多线程）"""
    seconds = duration_ms / 1000.0
    cmd = (
        f"ffmpeg -threads 0 -hide_banner -loglevel error -f lavfi -t {seconds} "
        f"-i anullsrc=r={sample_rate}:cl=mono -c:a aac -b:a 96k {shlex.quote(str(out_m4a))} -y"
    )
    run(cmd)


def concat_m4a(parts: list[Path], out_m4a: Path) -> None:
    """拼接 M4A 片段（编码统一为 AAC，多线程）"""
    with tempfile.TemporaryDirectory() as td:
        list_file = Path(td) / "list.txt"
        with open(list_file, "w", encoding="utf-8") as f:
            for p in parts:
                f.write(f"file '{p.as_posix()}'\n")
        cmd = f"ffmpeg -threads 0 -hide_banner -loglevel error -f concat -safe 0 -i {shlex.quote(str(list_file))} -c:a aac -b:a 192k {shlex.quote(str(out_m4a))} -y"
        run(cmd)


def enhance_with_white_noise_and_reverb(input_m4a: Path, white_noise_wav: Path, output_m4a: Path, noise_volume: float = 0.7) -> None:
    """使用 ffmpeg 为音频添加白噪音和房间混响效果（白噪音 70% 音量，多线程处理）"""
    # 白噪音 70%, 房间混响小空间, 动态压缩, EQ, 高通 80Hz, 响度归一
    filter_complex = (
        f"[0:a]volume=1.0[voice];"
        f"[1:a]volume={noise_volume:.2f},aloop=loop=-1:size=2e+09[noise];"
        f"[voice][noise]amix=inputs=2:duration=longest:weights=1:{noise_volume:.2f}[mixed];"
        f"[mixed]acompressor=threshold=-18:ratio=3:attack=15:release=180:makeup=3,"
        f"equalizer=f=250:width=120:g=2.0,"
        f"equalizer=f=3500:width=800:g=2.5,"
        f"highpass=f=80,"
        f"loudnorm=I=-19:TP=-2:LRA=9[output]"
    )
    # ffmpeg 多线程（-threads 0 自动检测 CPU 核心数）
    cmd = (
        f"ffmpeg -threads 0 -hide_banner -loglevel error "
        f"-i {shlex.quote(str(input_m4a))} "
        f"-i {shlex.quote(str(white_noise_wav))} "
        f"-filter_complex {shlex.quote(filter_complex)} "
        f"-map [output] -c:a aac -b:a 192k {shlex.quote(str(output_m4a))} -y"
    )
    run(cmd)


def find_column_index(header_row, target_names: list[str]) -> int | None:
    for idx, cell in enumerate(header_row, start=1):
        value = (cell.value or "").strip() if isinstance(cell.value, str) else str(cell.value or "").strip()
        if not value:
            continue
        if value in target_names:
            return idx
    return None


def process_workbook(
    xlsx_path: Path,
    out_dir: Path,
    voice: str,
    rate: str,
    volume: str,
    pitch: str,
    silence_ms: int,
) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    out_dir.mkdir(parents=True, exist_ok=True)

    # ffmpeg 增强版输出目录（20.2_ffpmeg输出文件_M4A格式音频文件）
    enhanced_base = out_dir.parent.parent / "20.2_ffpmeg输出文件_M4A格式音频文件"
    enhanced_base.mkdir(parents=True, exist_ok=True)

    # 白噪音文件路径
    script_dir = Path(__file__).parent
    white_noise_wav = script_dir / "white_noise.wav"
    possible_locations = [
        script_dir / "white_noise.wav",
        Path("/Volumes/M2/TT_Live_AI_TTS/white_noise.wav"),
        Path("/Volumes/M2/白噪音样本/white_noise.wav"),
    ]
    for loc in possible_locations:
        if loc.exists():
            white_noise_wav = loc
            break
    if not white_noise_wav.exists():
        print(f"[WARN] white_noise.wav not found, enhancement will be skipped. Please place it at {script_dir / 'white_noise.wav'}")
        white_noise_wav = None

    # Prebuild silence once per run to reuse (M4A format)
    silence_m4a = out_dir / f"__silence_{silence_ms}ms__.m4a"
    if not silence_m4a.exists():
        make_silence_m4a(silence_ms, silence_m4a)

    name_columns = ["文件名", "filename", "标题", "title", "名称", "name", "ID", "id", "序号", "index"]
    # 扩展英文文案列集合（含大小写/空格变体）
    english_columns = [
        "英文",
        "english_script", "English Script", "english", "English", "script", "Script",
        "english_text", "English Text", "Content", "content", "Text", "text",
        "Description", "description", "Copy", "copy", "Scripts", "scripts",
        "Prompts", "prompts", "Messages", "messages", "Posts", "posts",
        "Ads", "ads", "Marketing", "marketing", "Sales", "sales",
        "Copywriting", "copywriting", "Headlines", "headlines", "Taglines", "taglines",
        "Slogans", "slogans", "Captions", "captions", "Titles", "titles",
        "Subtitles", "subtitles", "Body", "body", "Main", "main",
        "Primary", "primary", "Core", "core", "Key", "key",
        "Essential", "essential", "Important", "important",
    ]

    # 情绪→基线参数映射
    emotion_baseline = {
        "Excited":  {"rate": 15.0,  "pitch": 12.0, "volume": 15.0},
        "Confident":{"rate": 8.0,   "pitch": 5.0,  "volume": 8.0},
        "Empathetic":{"rate": -12.0, "pitch": -8.0, "volume": -10.0},
        "Calm":     {"rate": -10.0, "pitch": -3.0, "volume": 0.0},
        "Playful":  {"rate": 18.0,  "pitch": 15.0, "volume": 5.0},
        "Urgent":   {"rate": 22.0,  "pitch": 8.0,  "volume": 18.0},
        "Authoritative": {"rate": 5.0,  "pitch": 3.0,  "volume": 10.0},
        "Friendly": {"rate": 12.0,  "pitch": 8.0,  "volume": 5.0},
        "Inspirational": {"rate": 10.0,  "pitch": 10.0, "volume": 12.0},
        "Serious":  {"rate": 0.0,   "pitch": 0.0,  "volume": 5.0},
        "Mysterious":{"rate": -8.0,  "pitch": 5.0,  "volume": -5.0},
        "Grateful": {"rate": 5.0,   "pitch": 8.0,  "volume": 8.0},
    }

    # 情绪→语音池（节选）
    emotion_voice_pool = {
        "Excited": ["en-US-AriaNeural", "en-US-EmmaNeural", "en-US-MichelleNeural"],
        "Confident": ["en-US-NancyNeural", "en-US-SerenaNeural", "en-US-BrandonNeural"],
        "Empathetic": ["en-US-AvaNeural", "en-US-JennyNeural", "en-US-AshleyNeural"],
        "Calm": ["en-US-DavisNeural", "en-US-AvaNeural", "en-US-JennyNeural"],
        "Playful": ["en-US-EmmaNeural", "en-US-AriaNeural", "en-US-FableNeural"],
        "Urgent": ["en-US-MichelleNeural", "en-US-NancyNeural", "en-US-BrandonNeural"],
        "Authoritative": ["en-US-SerenaNeural", "en-US-NancyNeural", "en-US-BrandonNeural"],
        "Friendly": ["en-US-JennyNeural", "en-US-AvaNeural", "en-US-KaiNeural"],
        "Inspirational": ["en-US-MichelleNeural", "en-US-BrandonNeural", "en-US-AriaNeural"],
        "Serious": ["en-US-SerenaNeural", "en-US-NancyNeural", "en-US-DavisNeural"],
        "Mysterious": ["en-US-DavisNeural", "en-US-SerenaNeural", "en-US-AvaNeural"],
        "Grateful": ["en-US-JennyNeural", "en-US-AvaNeural", "en-US-AshleyNeural"],
    }

    fallback_voices = ["en-US-JennyNeural", "en-US-AriaNeural", "en-US-DavisNeural"]

    # 产品名关键词→情绪（简化映射）
    keyword_emotion = [
        (re.compile(r"美白|淡斑|亮白|brightening", re.I), "Excited"),
        (re.compile(r"抗老|紧致|firming|anti-aging", re.I), "Confident"),
        (re.compile(r"保湿|补水|滋润|moisturizing", re.I), "Calm"),
        (re.compile(r"vitamin|精华|serum", re.I), "Playful"),
        (re.compile(r"collagen|健康|health", re.I), "Empathetic"),
        (re.compile(r"瘦身|减肥|fitness|weight", re.I), "Inspirational"),
        (re.compile(r"护发|hair|柔顺|smooth", re.I), "Soothing"),  # 未在12种中，回退映射
        (re.compile(r"眼部|eye|温和|gentle", re.I), "Gentle"),        # 未在12种中，回退映射
        (re.compile(r"限时|紧急|秒杀|urgent|limited|deadline", re.I), "Urgent"),
        (re.compile(r"公告|通知|声明|official|announcement", re.I), "Serious"),
        (re.compile(r"感谢|回馈|grateful|thank", re.I), "Grateful"),
        (re.compile(r"新品|促销|sale|promotion|new", re.I), "Excited"),
        (re.compile(r"科技|高端|专业|premium|luxury|professional", re.I), "Confident"),
        (re.compile(r"家居|教育|home|education|learning", re.I), "Calm"),
        (re.compile(r"美妆|时尚|makeup|fashion|style", re.I), "Playful"),
    ]

    def detect_emotion_from_name(name: str) -> str:
        for pattern, emo in keyword_emotion:
            if pattern.search(name):
                # 映射到12种
                if emo in emotion_baseline:
                    return emo
                # 简单回退
                return "Calm"
        return "Excited"  # 默认

    def clamp(v: float, lo: float, hi: float) -> float:
        return max(lo, min(hi, v))

    def fmt_rate(pct: float) -> str:
        pct = clamp(pct, -50.0, 200.0)
        sign = "+" if pct >= 0 else ""
        return f"{sign}{pct:.0f}%"

    def fmt_pitch(hz: float) -> str:
        hz = clamp(hz, -50.0, 50.0)
        sign = "+" if hz >= 0 else ""
        return f"{sign}{hz:.0f}Hz"

    def fmt_volume(pct: float) -> str:
        pct = clamp(pct, -50.0, 50.0)
        sign = "+" if pct >= 0 else ""
        return f"{sign}{pct:.0f}%"

    # 从 XLSX 文件名提取产品名称和日期
    product_name, date_str = extract_product_and_date(xlsx_path.name)
    
    total_rows = 0
    generated = 0

    for ws in wb.worksheets:
        header = ws[1]
        # 优先使用“英文”，否则使用扩展集合
        english_col = find_column_index(header, ["英文"]) or find_column_index(header, english_columns)
        if not english_col:
            continue
        name_col = find_column_index(header, name_columns)

        # 所有工作表的内容都放在同一个 xlsx 文件夹，不分子文件夹
        # 文件名添加工作表前缀避免重名
        sheet_prefix = sanitize_filename(ws.title)

        for row_idx in range(2, ws.max_row + 1):
            total_rows += 1
            text_cell = ws.cell(row=row_idx, column=english_col)
            text = text_cell.value if isinstance(text_cell.value, str) else str(text_cell.value or "")
            text = text.strip()
            if not text:
                continue

            # 文件名格式：产品名_日期_工作表名_原始文件名
            base_name: str
            if name_col:
                name_val = ws.cell(row=row_idx, column=name_col).value
                name_val = name_val if isinstance(name_val, str) else str(name_val or "")
                if name_val.strip():
                    original_name = sanitize_filename(name_val)
                else:
                    original_name = f"row_{row_idx}"
            else:
                original_name = f"row_{row_idx}"
            
            # 组合完整文件名：产品名_日期_工作表名_原始文件名
            base_name = f"{product_name}_{date_str}_{sheet_prefix}_{original_name}"

            out_m4a = out_dir / f"{base_name}.m4a"
            if out_m4a.exists():
                # Skip existing
                continue

            sentences = split_into_sentences(text)

            # 从文件名和工作表名推断情绪
            product_hint = f"{xlsx_path.stem}_{ws.title}"
            emotion = detect_emotion_from_name(product_hint)
            baseline = emotion_baseline.get(emotion, {"rate": 0.0, "pitch": 0.0, "volume": 0.0})
            voice_pool = emotion_voice_pool.get(emotion, fallback_voices)

            # 确定“本文件固定使用的 voice”
            # 优先使用命令行传入的 --voice；若传入的是默认占位（如 "auto" 或空），则按文件名哈希在情绪池中选择一个
            def pick_file_voice() -> str:
                v = voice
                if v and v.lower() not in ("auto", "none"):
                    return v
                # 根据文件名确定性选择，保证同一 xlsx 固定同一 voice
                key = abs(hash(xlsx_path.stem))
                pool = voice_pool if voice_pool else fallback_voices
                return pool[key % len(pool)]

            file_voice = pick_file_voice()

            # 按情绪微调句间静音
            silence_adjust = 0
            if emotion in ("Calm", "Empathetic"):
                silence_adjust = 150
            elif emotion == "Urgent":
                silence_adjust = -200
            row_silence_ms = max(200, silence_ms + silence_adjust)

            # 多线程 TTS 生成（批量异步并行处理所有句子）
            async def generate_segments_parallel():
                async_tasks = []
                seg_info = []
                for i, sent in enumerate(sentences):
                    seg_m4a = td_path / f"seg_{i:03d}.m4a"
                    rate_val = baseline["rate"] + math.sin(i * math.pi / 6.0) * 2.0
                    dyn_rate = fmt_rate(rate_val)
                    seg_info.append((i, seg_m4a))
                    async_tasks.append(tts_segment_async(sent, file_voice, dyn_rate, seg_m4a))
                
                # 并行执行所有任务
                results = await asyncio.gather(*async_tasks, return_exceptions=True)
                
                # 组装结果
                final_results = []
                for idx, (result, (seg_idx, seg_m4a)) in enumerate(zip(results, seg_info)):
                    if isinstance(result, Exception):
                        print(f"[WARN] TTS failed for row {row_idx} sent {seg_idx}: {result}")
                        final_results.append((seg_idx, seg_m4a, False))
                    else:
                        final_results.append((seg_idx, seg_m4a, True))
                return final_results

            parts: list[Path] = []
            with tempfile.TemporaryDirectory() as td:
                td_path = Path(td)
                # 并行生成所有句子片段
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    segment_results = loop.run_until_complete(generate_segments_parallel())
                finally:
                    loop.close()

                # 按顺序组装片段和静音
                for i, (seg_idx, seg_m4a, success) in enumerate(sorted(segment_results)):
                    if success and seg_m4a.exists():
                        parts.append(seg_m4a)
                    if i < len(sentences) - 1 and row_silence_ms > 0:
                        cache_sil = out_dir / f"__silence_{row_silence_ms}ms__.m4a"
                        if not cache_sil.exists():
                            make_silence_m4a(row_silence_ms, cache_sil)
                        parts.append(cache_sil)

                if parts:
                    concat_m4a(parts, out_m4a)

            generated += 1
            print(f"[{xlsx_path.name}][{ws.title}] row {row_idx} [{emotion}] -> {out_m4a}")

            # 增强处理：添加白噪音和房间声效（同样直接放在 xlsx 文件夹，不分子文件夹）
            if white_noise_wav and white_noise_wav.exists():
                # 增强版也放在同一个 xlsx 文件夹
                enhanced_dir = enhanced_base / xlsx_path.stem
                enhanced_dir.mkdir(parents=True, exist_ok=True)
                enhanced_m4a = enhanced_dir / f"{base_name}.m4a"
                if not enhanced_m4a.exists():
                    try:
                        enhance_with_white_noise_and_reverb(out_m4a, white_noise_wav, enhanced_m4a, noise_volume=0.7)
                        print(f"  -> Enhanced: {enhanced_m4a}")
                    except Exception as e:
                        print(f"  [WARN] Enhance failed for {out_m4a}: {e}")

    print(f"Done: {xlsx_path}  total_rows={total_rows}  generated={generated}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate TTS M4A from XLSX '英文' column using edge-tts (multi-threaded) and ffmpeg (multi-threaded)")
    parser.add_argument("xlsx", type=str, help="Path to XLSX file")
    parser.add_argument("--out", type=str, default="out_audio", help="Output directory root")
    parser.add_argument("--voice", type=str, default="en-US-JennyNeural", help="edge-tts voice")
    parser.add_argument("--rate", type=str, default="0%", help="edge-tts rate, e.g., -25%% or 0%%")
    parser.add_argument("--volume", type=str, default="0%", help="edge-tts volume, e.g., -10%% or 0%%")
    parser.add_argument("--pitch", type=str, default="0Hz", help="edge-tts pitch, e.g., -50Hz or 0Hz")
    parser.add_argument("--silence-ms", type=int, default=600, help="Silence duration between sentences in ms")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    out_root = Path(args.out)
    xlsx_out = out_root / xlsx_path.stem
    xlsx_out.mkdir(parents=True, exist_ok=True)

    process_workbook(
        xlsx_path=xlsx_path,
        out_dir=xlsx_out,
        voice=args.voice,
        rate=args.rate,
        volume=args.volume,
        pitch=args.pitch,
        silence_ms=args.silence_ms,
    )


if __name__ == "__main__":
    main()


